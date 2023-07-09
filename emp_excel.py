import openpyxl
from openpyxl import Workbook,load_workbook

FILE_PATH = r"C:\Users\vinit\OneDrive\Documents\Softskill\Course python\Practice\Excel Handling\emp_records.xlsx"
FILE_PATH_2 = r"C:\Users\vinit\OneDrive\Documents\Softskill\Course python\Practice\Excel Handling\emp_excel_csv.csv"
# wb= Workbook()
# wb.create_sheet("Employees")
# del wb["Sheet"]
# wb.save(FILE_PATH)

def load_excel_file():
    wb = load_workbook(FILE_PATH)
    sheet = wb.active
    return wb, sheet
    